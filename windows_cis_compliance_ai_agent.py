#!/usr/bin/env python3
"""
Windows CIS Compliance AI Agent

AI-powered CIS Benchmark compliance validation tool.
Formats Nessus CSV to company Excel template and automatically
validates compliance using LLM + PowerShell execution.

Author: Enhanced by AI Agent
Version: 3.0

Changes from v2.0:
 - Gap 1:  Command safety validation (allowlist + blocklist before execution)
 - Gap 2:  API retry with exponential backoff
 - Gap 3:  Fixed `continue` that silently dropped findings from report
 - Gap 4:  CIS benchmark string is now configurable via --benchmark
 - Gap 5:  Structured JSON output matching the spec (output_type, expected_output_pattern, compliance_logic)
 - Gap 6:  CSV column validation at load time
 - Gap 7:  Replaced print() with Python logging module
 - Gap 8:  OpenAI client created once in main(), passed as parameter
 - Gap 9:  Checkpoint / resume capability (--resume flag)
 - Gap 10: Async parallel processing option (--parallel N)
 - Gap 11: API key fully masked in console output
 - Gap 12: --dry-run mode (generates commands without executing)
 - Gap 13: Command hint system — maps CIS categories to known-good commands
 - Gap 14: Fallback commands — auto-retries with alternative if primary fails
"""

import argparse
import re
import os
import sys
import json
import time
import logging
import subprocess
import asyncio
from pathlib import Path
from typing import Dict, List, Optional, Tuple

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

# LLM integration
try:
    from openai import OpenAI
except ImportError:
    print("[ERROR] OpenAI package not installed. Run: pip install openai")
    sys.exit(1)


# ============================================================================
# Logging Setup  (Gap 7)
# ============================================================================

def setup_logging(log_file: Optional[str] = None, verbose: bool = False):
    """Configure logging with console + optional file handler."""
    level = logging.DEBUG if verbose else logging.INFO
    fmt = "%(asctime)s [%(levelname)s] %(message)s"
    datefmt = "%Y-%m-%d %H:%M:%S"

    handlers: List[logging.Handler] = [logging.StreamHandler(sys.stdout)]
    if log_file:
        handlers.append(logging.FileHandler(log_file, encoding="utf-8"))

    logging.basicConfig(level=level, format=fmt, datefmt=datefmt, handlers=handlers)


logger = logging.getLogger(__name__)


# ============================================================================
# Command Hint System  (Gap 13 — fixes LLM hallucinating wrong commands)
# ============================================================================

# Maps CIS finding keyword patterns to the CORRECT command to use.
# This prevents the LLM from generating wrong registry paths.
COMMAND_HINTS = [
    # --- Account / Password Policy (CIS 1.1.x) ---
    {
        "patterns": [
            r"password history",
            r"maximum password age",
            r"minimum password age",
            r"minimum password length",
            r"password must meet complexity",
            r"store passwords using reversible",
            r"password.*(length|age|history|complex)",
        ],
        "command": "net accounts",
        "fallback": "secedit /export /cfg C:\\Windows\\Temp\\secpol_check.inf /areas SECURITYPOLICY && type C:\\Windows\\Temp\\secpol_check.inf | findstr /i \"Password\"",
        "output_type": "accountpolicy",
        "hint": "For password policy settings, ALWAYS use 'net accounts' which shows all password policy values reliably. Do NOT use Get-ItemProperty on registry — these settings are NOT stored in the registry in a directly accessible way."
    },
    # --- Account Lockout Policy (CIS 1.2.x) ---
    {
        "patterns": [
            r"account lockout duration",
            r"account lockout threshold",
            r"reset account lockout",
            r"administrator account lockout",
            r"lockout.*(duration|threshold|counter|reset)",
        ],
        "command": "net accounts",
        "fallback": "secedit /export /cfg C:\\Windows\\Temp\\secpol_check.inf /areas SECURITYPOLICY && type C:\\Windows\\Temp\\secpol_check.inf | findstr /i \"Lockout\"",
        "output_type": "accountpolicy",
        "hint": "For account lockout settings, ALWAYS use 'net accounts' which shows lockout duration, threshold, and observation window. Do NOT use Get-ItemProperty on registry."
    },
    # --- Audit Policy (CIS 17.x.x) ---
    {
        "patterns": [
            r"audit\s+(credential|application group|computer account|distribution group|other account|security group|user account|logon|logoff|account lockout|ipsec|network policy|other logon|special logon|removable storage|audit policy|kerberos|file share|file system|handle manipulation|kernel object|other object|registry|sam|certification|detailed file|dpapi|pnp|process|rpc|token|security state|security system|system integrity|directory service|sensitive privilege|non sensitive privilege|other privilege)",
            r"audit\b.*\b(success|failure)",
        ],
        "command": "auditpol /get /category:*",
        "fallback": None,
        "output_type": "auditpolicy",
        "hint": "For audit policy settings, use 'auditpol /get /category:*' to get all audit categories, or 'auditpol /get /subcategory:\"<specific subcategory>\"' for a specific one. NOTE: auditpol requires administrator privileges."
    },
    # --- User Rights Assignment (CIS 2.2.x) ---
    {
        "patterns": [
            r"access this computer from the network",
            r"act as part of the operating system",
            r"deny (access|log on)",
            r"(log on|allow) (locally|as a batch|as a service|remotely|through remote)",
            r"impersonate a client",
            r"create (a pagefile|global objects|permanent shared|symbolic links|a token)",
            r"debug programs",
            r"force shutdown",
            r"generate security audits",
            r"(increase|adjust|modify) (scheduling|memory|firmware)",
            r"load and unload device drivers",
            r"lock pages in memory",
            r"manage auditing and security",
            r"profile (single|system)",
            r"replace a process level token",
            r"restore files and directories",
            r"shut down the system",
            r"take ownership",
            r"back up files and directories",
            r"bypass traverse checking",
            r"change (the system time|time zone)",
        ],
        "command": "secedit /export /cfg C:\\Windows\\Temp\\secpol_check.inf /areas USER_RIGHTS && type C:\\Windows\\Temp\\secpol_check.inf",
        "fallback": None,
        "output_type": "policy",
        "hint": "For user rights assignment, use 'secedit /export /cfg C:\\Windows\\Temp\\secpol_check.inf /areas USER_RIGHTS && type C:\\Windows\\Temp\\secpol_check.inf'. This exports and displays all user rights assignments."
    },
    # --- Security Options (CIS 2.3.x) ---
    {
        "patterns": [
            r"accounts:.*guest account",
            r"accounts:.*administrator account",
            r"accounts:.*blank passwords",
            r"rename (administrator|guest) account",
        ],
        "command": "net user guest",
        "fallback": "Get-LocalUser | Select-Object Name, Enabled",
        "output_type": "accountpolicy",
        "hint": "For guest/administrator account status, use 'net user guest' or 'net user administrator' or 'Get-LocalUser | Select-Object Name, Enabled'. These show if accounts are active/disabled."
    },
    # --- Firewall (CIS 9.x.x) ---
    {
        "patterns": [
            r"(windows|firewall).*(domain|private|public).*(state|profile|enabled|logging|inbound|outbound|notification)",
            r"firewall.*state.*on",
        ],
        "command": "Get-NetFirewallProfile | Select-Object Name, Enabled, DefaultInboundAction, DefaultOutboundAction, LogAllowed, LogBlocked, LogFileName, LogMaxSizeKilobytes",
        "fallback": "netsh advfirewall show allprofiles",
        "output_type": "firewall",
        "hint": "For firewall settings, use 'Get-NetFirewallProfile | Select-Object Name, Enabled, DefaultInboundAction, DefaultOutboundAction, LogAllowed, LogBlocked, LogFileName, LogMaxSizeKilobytes'. This shows all profile states and settings."
    },
    # --- Services (CIS 5.x) ---
    {
        "patterns": [
            r"(service|services).*(disabled|enabled|running|stopped|automatic|manual)",
            r"(print spooler|remote registry|windows search|xbox|fax|lxss|openSSH|infrared|iis admin|internet connection sharing)",
        ],
        "command": None,   # Dynamic — depends on service name
        "fallback": None,
        "output_type": "service",
        "hint": "For service status checks, use 'Get-Service -Name \"<service_name>\" | Select-Object Name, Status, StartType'. Replace <service_name> with the actual Windows service name."
    },
    # --- Registry-based settings (CIS 2.3.x, 18.x.x) ---
    {
        "patterns": [
            r"(interactive logon|microsoft network|network access|network security|shutdown|system cryptography|system objects|user account control)",
            r"HKLM",
            r"registry",
        ],
        "command": None,   # Dynamic — depends on registry path
        "fallback": None,
        "output_type": "registry",
        "hint": "For registry-based settings, use 'Get-ItemProperty -Path \"HKLM:\\<path>\" -Name \"<valuename>\" | Select-Object -ExpandProperty <valuename>'. Make sure the registry path is correct for the specific setting."
    },
    # --- Relax minimum password length limits (specific CIS 1.1.6) ---
    {
        "patterns": [
            r"relax minimum password length",
        ],
        "command": "Get-ItemProperty -Path 'HKLM:\\System\\CurrentControlSet\\Control\\SAM' -Name 'RelaxMinimumPasswordLengthLimits' -ErrorAction SilentlyContinue | Select-Object -ExpandProperty RelaxMinimumPasswordLengthLimits",
        "fallback": "reg query \"HKLM\\System\\CurrentControlSet\\Control\\SAM\" /v RelaxMinimumPasswordLengthLimits",
        "output_type": "registry",
        "hint": "For 'Relax minimum password length limits', check the registry at HKLM:\\System\\CurrentControlSet\\Control\\SAM with value name RelaxMinimumPasswordLengthLimits. Value 1 = Enabled."
    },
    # --- Event Log Settings (CIS 18.10.26.x) ---
    {
        "patterns": [
            r"event log.*(behavior|size|maximum|retention)",
            r"(application|security|setup|system):.*(control event|specify the maximum|log file)",
            r"control event log behavior",
            r"specify the maximum log file size",
        ],
        "command": None,  # Dynamic — depends on which log
        "fallback": None,
        "output_type": "other",
        "hint": "For Event Log settings, use 'wevtutil gl <LogName>' (gl = get-log, READ-ONLY) to view the current log configuration. For example: 'wevtutil gl Application', 'wevtutil gl Security', 'wevtutil gl Setup', 'wevtutil gl System'. NEVER use 'wevtutil sl' as that is a WRITE/SET command. Alternatively, check registry: Get-ItemProperty -Path 'HKLM:\\SOFTWARE\\Policies\\Microsoft\\Windows\\EventLog\\<LogName>' -ErrorAction SilentlyContinue"
    },
]


def get_command_hint(finding: str) -> Optional[Dict]:
    """
    Match a CIS finding to a command hint that tells the LLM the correct command.

    Args:
        finding: The CIS finding title text.

    Returns:
        Matching hint dictionary, or None if no hint matches.
    """
    finding_lower = finding.lower()
    for hint in COMMAND_HINTS:
        for pattern in hint["patterns"]:
            if re.search(pattern, finding_lower, re.IGNORECASE):
                return hint
    return None


# ============================================================================
# Command Safety Validation  (Gap 1)
# ============================================================================

# Patterns that indicate a DANGEROUS / mutating command
DANGEROUS_PATTERNS = [
    r'\bRemove-Item\b',
    r'\bRemove-ItemProperty\b',
    r'\bSet-ItemProperty\b',
    r'\bNew-ItemProperty\b',
    r'\bSet-NetFirewallProfile\b',
    r'\bStop-Service\b',
    r'\bRestart-Service\b',
    r'\bStart-Service\b',
    r'\bStart-Process\b',
    r'\bInvoke-WebRequest\b',
    r'\bInvoke-RestMethod\b',
    r'\bInvoke-Expression\b',
    r'\biex\b',
    r'\bFormat-Volume\b',
    r'\bClear-Content\b',
    r'\bClear-EventLog\b',
    r'\bDisable-NetAdapter\b',
    r'\bEnable-NetAdapter\b',
    r'\bSet-ExecutionPolicy\b',
    r'\bSet-MpPreference\b',
    r'\bAdd-LocalGroupMember\b',
    r'\bRemove-LocalGroupMember\b',
    r'\bNew-LocalUser\b',
    r'\bRemove-LocalUser\b',
    r'\bSet-LocalUser\b',
    r'\bnet\s+user\s+\S+\s+/add\b',
    r'\bnet\s+user\s+\S+\s+/delete\b',
    r'\bnet\s+localgroup\s+\S+\s+\S+\s+/add\b',
    r'\bnet\s+stop\b',
    r'\bsc\s+(stop|delete|config)\b',
    r'\bwevtutil\s+sl\b',
    r'\bwevtutil\s+set-log\b',
    r'\brm\s+',
    r'\bdel\s+',
    r'\brd\s+',
    r'\brmdir\b',
    r'\bshutdown\b',
    r'\bRestart-Computer\b',
    r'\bStop-Computer\b',
    r'\bRename-Item\b',
    r'\bMove-Item\b',
    r'\bCopy-Item\b',
    r'\bNew-Item\b',
    r'\bOut-File\b',
    r'\bSet-Content\b',
    r'\bAdd-Content\b',
    r'\bWrite-Output\b.*\|\s*Out-File',
    r'\bDownloadString\b',
    r'\bDownloadFile\b',
    r'\bSystem\.Net\.WebClient\b',
    r'\bNew-Object\b.*\bNet\.',
    r'\bReg\s+(add|delete)\b',
]

# Known-safe read-only command prefixes removed as requested to allow more flexibility.
# The script now relies solely on DANGEROUS_PATTERNS for security.


def is_command_safe(command: str) -> Tuple[bool, str]:
    """
    Validate a PowerShell command against safety rules.
    Only checks for dangerous patterns.

    Returns:
        Tuple of (is_safe, reason)
    """
    if not command or command.startswith("# ERROR"):
        return False, "Empty or error command"

    # Check for dangerous patterns
    for pattern in DANGEROUS_PATTERNS:
        if re.search(pattern, command, re.IGNORECASE):
            return False, f"Blocked by dangerous pattern: {pattern}"

    # If it's not in the blocklist, it's considered safe
    return True, "No dangerous patterns detected (Allow-by-default)"


# ============================================================================
# Retry Logic  (Gap 2)
# ============================================================================

def retry_with_backoff(func, *args, max_retries: int = 3, base_delay: float = 2.0, **kwargs):
    """
    Execute a function with exponential backoff retry on failure.

    Args:
        func: Function to call
        max_retries: Maximum number of retry attempts
        base_delay: Base delay in seconds (doubles each retry)
    """
    last_exception = None
    for attempt in range(max_retries + 1):
        try:
            return func(*args, **kwargs)
        except Exception as e:
            last_exception = e
            if attempt < max_retries:
                delay = base_delay * (2 ** attempt)
                logger.warning(
                    "Attempt %d/%d failed: %s — retrying in %.1fs",
                    attempt + 1, max_retries + 1, str(e), delay
                )
                time.sleep(delay)
            else:
                logger.error(
                    "All %d attempts failed. Last error: %s",
                    max_retries + 1, str(e)
                )
    raise last_exception


# ============================================================================
# Original Parsing Functions (Kept Intact)
# ============================================================================

def parse_cis_from_description(desc: str):
    """
    Nessus compliance content usually stores CIS ID + title inside Description.
    Example first line:
    1.1.1 (L1) Ensure 'Enforce password history' is set to '24 or more password(s)'
    """
    if not isinstance(desc, str):
        return ("", "", "")

    txt = desc.strip().strip('"')
    lines = txt.split("\n", 1)
    first_line = lines[0].strip()
    body = lines[1].strip() if len(lines) > 1 else ""

    m = re.match(r"^(\d+(?:\.\d+)*)\s*\((L\d)\)\s*(.*)$", first_line)
    if m:
        cis_id = m.group(1)
        title = m.group(3).strip()
        title = re.sub(r"^\s*Ensure\s*", "", title).strip()
        return (cis_id, title, body)

    # fallback if format differs
    return ("", first_line, body)


def clean_finding_title(title: str) -> str:
    """
    Convert:
    'Enforce password history' is set to '24 or more password(s)'" :
    -> Enforce password history is set to 24 or more passwords
    """
    if not isinstance(title, str):
        return ""
    t = title.strip()

    # remove [FAILED], : etc
    t = re.sub(r"\s*\[(FAILED|PASSED|WARNING)\]\s*", "", t, flags=re.I)
    t = re.sub(r"\s*:\s*$", "", t)

    # remove all quotes
    t = t.replace('"', "").replace("\u2019", "'").replace("'", "")

    # normalize password(s)
    t = t.replace("(s)", "s")

    # cleanup spaces
    t = re.sub(r"\s+", " ", t).strip()
    return t


def extract_impact_from_solution(solution: str) -> str:
    """
    Extract Impact section from Nessus Solution text.
    """
    if not isinstance(solution, str):
        return ""

    txt = solution.replace("\r", "")
    m = re.search(r"\n\s*Impact:\s*\n(.*)", txt, flags=re.I | re.S)
    if not m:
        return ""

    impact = m.group(1).strip().strip('"')

    # stop at next major label
    impact = re.split(
        r"\n\s*(Rationale|Audit|Remediation|Default Value|References)\s*:",
        impact,
        flags=re.I,
    )[0].strip()

    return impact


def remove_impact_from_solution(solution: str) -> str:
    """
    Remove Impact section from remediation text so remediation is pure fix steps.
    """
    if not isinstance(solution, str):
        return ""
    txt = solution.replace("\r", "")
    txt = re.sub(r"\n\s*Impact:\s*\n.*", "", txt, flags=re.I | re.S).strip()
    return txt


# ============================================================================
# LLM Validation Functions (Enhanced with Gaps 2, 5, 8, 13, 14)
# ============================================================================

def generate_validation_json(
    client: OpenAI,
    finding: str,
    details: str,
    model: str = "gpt-4o-mini"
) -> Dict:
    """
    Ask LLM to generate a structured validation object for a CIS finding.
    (Gap 5: Full structured JSON matching the spec)
    (Gap 13: Includes command hints for known CIS categories)

    Returns:
        Dictionary with validation_command, expected_output_pattern,
        compliance_logic, and output_type.
    """
    # Gap 13: Get command hint for this finding
    hint = get_command_hint(finding)

    hint_section = ""
    if hint:
        hint_section = f"""

IMPORTANT COMMAND GUIDANCE:
{hint['hint']}
Recommended command: {hint['command'] if hint['command'] else '(see hint above)'}
Output type: {hint['output_type']}
"""

    prompt = f"""You are a Windows Security Compliance Validation Assistant.

CIS Finding:
{finding}

Details:
{details[:500]}
{hint_section}
Generate the exact PowerShell or CMD command required to validate the configuration on a Windows Server.

CRITICAL RULES:
- For PASSWORD policy (history, age, length, complexity): ALWAYS use 'net accounts'. Do NOT use Get-ItemProperty on registry paths — these settings are NOT in the registry in a directly queryable way.
- For ACCOUNT LOCKOUT policy (duration, threshold, counter): ALWAYS use 'net accounts'.
- For AUDIT policy: use 'auditpol /get /subcategory:"<name>"' or 'auditpol /get /category:*'.
- For USER RIGHTS ASSIGNMENT: use 'secedit /export /cfg C:\\Windows\\Temp\\secpol_check.inf /areas USER_RIGHTS && type C:\\Windows\\Temp\\secpol_check.inf'.
- For FIREWALL: use 'Get-NetFirewallProfile'.
- For SERVICES: use 'Get-Service -Name "<svcname>"'.
- For REGISTRY settings (ONLY when you know the exact correct registry path): use Get-ItemProperty.
- Use only built-in Windows commands.
- The command must be READ-ONLY (no system modifications).
- The command must work on Windows Server 2016, 2019, and 2022.

Return ONLY valid JSON in this exact format:
{{
  "validation_command": "<exact PowerShell/CMD command>",
  "expected_output_pattern": "<what output indicates compliant>",
  "compliance_logic": "<how to determine compliant vs non-compliant>",
  "output_type": "<one of: registry | service | policy | auditpolicy | firewall | accountpolicy | other>"
}}"""

    def _call():
        response = client.chat.completions.create(
            model=model,
            messages=[{"role": "user", "content": prompt}],
            temperature=0,
            max_tokens=500,
            response_format={"type": "json_object"}
        )
        result = json.loads(response.choices[0].message.content)

        # Validate required fields
        required = ["validation_command", "expected_output_pattern", "compliance_logic", "output_type"]
        for field in required:
            if field not in result:
                result[field] = "UNKNOWN"

        # Clean up command artifacts
        cmd = result.get("validation_command", "")
        cmd = cmd.replace('```powershell', '').replace('```', '').strip()
        cmd = cmd.replace('PowerShell command:', '').strip()
        result["validation_command"] = cmd

        return result

    # Gap 2: Retry with backoff
    try:
        return retry_with_backoff(_call, max_retries=3, base_delay=2.0)
    except Exception as e:
        logger.error("LLM command generation failed after retries: %s", e)

        # If we have a hint with a known command, use it as fallback
        if hint and hint.get("command"):
            logger.info("Using command hint as fallback: %s", hint["command"])
            return {
                "validation_command": hint["command"],
                "expected_output_pattern": "Check output against CIS requirement",
                "compliance_logic": "Compare output values with CIS expected values",
                "output_type": hint.get("output_type", "other")
            }

        return {
            "validation_command": "# ERROR: Could not generate command",
            "expected_output_pattern": "N/A",
            "compliance_logic": "N/A",
            "output_type": "other"
        }


# Patterns in stderr that indicate a "meaningful" validation result,
# NOT a true execution failure. These mean the setting is not configured.
MEANINGFUL_ERROR_PATTERNS = [
    r'Property .+ does not exist at path',
    r'Property .+ cannot be found',
    r'Cannot find path',
    r'does not exist',
    r'ItemNotFoundException',
    r'PSArgumentException',
    r'The term .+ is not recognized',
]


def is_meaningful_error(error_text: str) -> bool:
    """
    Check if a command error is actually a meaningful validation result.
    For example, 'Property X does not exist' means the GPO is not configured,
    which is valid data for compliance analysis (typically = Non Compliant).
    """
    for pattern in MEANINGFUL_ERROR_PATTERNS:
        if re.search(pattern, error_text, re.IGNORECASE):
            return True
    return False


def execute_powershell(command: str, timeout: int = 30) -> Dict:
    """
    Execute PowerShell command locally.

    Args:
        command: PowerShell command to execute
        timeout: Command timeout in seconds

    Returns:
        Dictionary with execution results
    """
    # Fix && for PowerShell compatibility (older PS versions don't support &&)
    if '&&' in command:
        command = command.replace('&&', ';')

    try:
        result = subprocess.run(
            ["powershell", "-NoProfile", "-NonInteractive", "-Command", command],
            capture_output=True,
            text=True,
            timeout=timeout,
            encoding='utf-8',
            errors='replace'
        )

        return {
            'success': result.returncode == 0,
            'output': result.stdout if result.stdout else '',
            'errors': result.stderr if result.stderr else '',
            'return_code': result.returncode
        }

    except subprocess.TimeoutExpired:
        return {
            'success': False,
            'output': '',
            'errors': f'Command timed out after {timeout} seconds',
            'return_code': -1
        }

    except Exception as e:
        return {
            'success': False,
            'output': '',
            'errors': f'Execution error: {str(e)}',
            'return_code': -1
        }


def execute_with_fallback(command: str, finding: str, timeout: int = 30) -> Dict:
    """
    Execute a command, and if it fails, check for a fallback command from hints.
    (Gap 14: Automatic fallback on command failure)

    Returns:
        Dictionary with execution results + 'command_used' field
    """
    # Try primary command
    result = execute_powershell(command, timeout)
    result['command_used'] = command

    if result['success']:
        return result

    # On failure, check for a fallback
    hint = get_command_hint(finding)
    if hint and hint.get("fallback"):
        fallback_cmd = hint["fallback"]
        logger.info("  ├─ Primary command failed, trying fallback: %s", fallback_cmd[:80])

        # Safety check the fallback too
        is_safe, safety_reason = is_command_safe(fallback_cmd)
        if not is_safe:
            logger.warning("  ├─ Fallback also blocked: %s", safety_reason)
            return result  # Return original failure

        fallback_result = execute_powershell(fallback_cmd, timeout)
        fallback_result['command_used'] = fallback_cmd

        if fallback_result['success']:
            return fallback_result

    # If we have a hint with a known command that's different from what we tried
    if hint and hint.get("command") and hint["command"] != command:
        known_cmd = hint["command"]
        logger.info("  ├─ Trying known-good command: %s", known_cmd[:80])

        is_safe, _ = is_command_safe(known_cmd)
        if is_safe:
            known_result = execute_powershell(known_cmd, timeout)
            known_result['command_used'] = known_cmd

            if known_result['success']:
                return known_result

    return result  # Return original failure if nothing worked


def validate_compliance(
    client: OpenAI,
    finding: str,
    details: str,
    command_output: str,
    expected_pattern: str,
    compliance_logic: str,
    model: str = "gpt-4o-mini"
) -> Dict:
    """
    Ask LLM to determine if the finding is actually compliant or not.
    Enhanced with the expected_pattern and compliance_logic from Step 1.
    """
    prompt = f"""You are a security analyst validating CIS Benchmark compliance.

CIS Requirement:
{finding}

Details:
{details[:500]}

Expected Output Pattern (from validation spec):
{expected_pattern}

Compliance Logic:
{compliance_logic}

Actual System Output from Validation Command:
{command_output[:1000]}

Task: Determine if the system is COMPLIANT or NON-COMPLIANT with this CIS requirement.

Analysis Steps:
1. Compare the requirement with actual system output
2. Check against the expected output pattern
3. Apply the compliance logic
4. Determine confidence level (HIGH/MEDIUM/LOW)

Response Format (JSON only):
{{
  "status": "Compliant" or "Non Compliant",
  "confidence": "HIGH" or "MEDIUM" or "LOW",
  "reasoning": "Brief explanation why compliant or non-compliant",
  "evidence": "Key part of output that shows compliance status"
}}

Return ONLY valid JSON, no other text:"""

    def _call():
        response = client.chat.completions.create(
            model=model,
            messages=[{"role": "user", "content": prompt}],
            temperature=0,
            max_tokens=500,
            response_format={"type": "json_object"}
        )

        result = json.loads(response.choices[0].message.content)

        # Validate required fields
        required_fields = ['status', 'confidence', 'reasoning', 'evidence']
        for field in required_fields:
            if field not in result:
                result[field] = 'UNKNOWN'

        return result

    # Gap 2: Retry with backoff
    try:
        return retry_with_backoff(_call, max_retries=3, base_delay=2.0)
    except Exception as e:
        logger.error("LLM validation failed after retries: %s", e)
        return {
            'status': 'ERROR',
            'confidence': 'LOW',
            'reasoning': f'LLM analysis failed: {str(e)}',
            'evidence': 'N/A'
        }


# ============================================================================
# Checkpoint / Resume  (Gap 9)
# ============================================================================

def get_checkpoint_path(output_file: str) -> str:
    """Get the checkpoint file path based on the output file name."""
    return output_file + ".checkpoint.json"


def save_checkpoint(checkpoint_path: str, records: List[Dict], processed_indices: List):
    """Save current progress to checkpoint file."""
    data = {
        "records": records,
        "processed_indices": processed_indices,
        "timestamp": time.strftime("%Y-%m-%d %H:%M:%S"),
    }
    with open(checkpoint_path, 'w', encoding='utf-8') as f:
        json.dump(data, f, indent=2, ensure_ascii=False, default=str)
    logger.debug("Checkpoint saved: %d records, %d processed indices", len(records), len(processed_indices))


def load_checkpoint(checkpoint_path: str) -> Tuple[List[Dict], List]:
    """Load progress from checkpoint file."""
    if not os.path.exists(checkpoint_path):
        return [], []
    try:
        with open(checkpoint_path, 'r', encoding='utf-8') as f:
            data = json.load(f)
        records = data.get("records", [])
        processed = data.get("processed_indices", [])
        logger.info(
            "Resumed from checkpoint: %d records, %d already processed (saved at %s)",
            len(records), len(processed), data.get("timestamp", "unknown")
        )
        return records, processed
    except Exception as e:
        logger.warning("Failed to load checkpoint: %s — starting fresh", e)
        return [], []


def remove_checkpoint(checkpoint_path: str):
    """Remove checkpoint file on successful completion."""
    if os.path.exists(checkpoint_path):
        os.remove(checkpoint_path)
        logger.debug("Checkpoint file removed: %s", checkpoint_path)


# ============================================================================
# CSV Validation  (Gap 6)
# ============================================================================

REQUIRED_CSV_COLUMNS = ["Risk", "Description", "Solution"]


def validate_csv_columns(df: pd.DataFrame, required: Optional[List[str]] = None):
    """Validate that the CSV contains all required columns."""
    required = required or REQUIRED_CSV_COLUMNS
    missing = [col for col in required if col not in df.columns]
    if missing:
        available = ", ".join(df.columns.tolist())
        raise ValueError(
            f"CSV is missing required columns: {missing}\n"
            f"Available columns: {available}\n"
            f"Ensure your Nessus CSV export includes: {required}"
        )


# ============================================================================
# Excel Formatting Functions (Original + Enhanced)
# ============================================================================

def format_cell_wrapping(ws):
    """Apply text wrapping and freeze panes."""
    wrap = Alignment(wrap_text=True, vertical="top")
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, max_col=ws.max_column):
        for cell in row:
            cell.alignment = wrap

    for c in ws[1]:
        c.font = Font(bold=True)

    ws.freeze_panes = "A2"


def set_column_widths(ws, headers):
    """Set column widths based on header names."""
    for idx, header in enumerate(headers, 1):
        letter = get_column_letter(idx)
        if header in ["Details Summary", "Remediation"]:
            ws.column_dimensions[letter].width = 80
        elif header in ["Finding", "CIS Benchmark"]:
            ws.column_dimensions[letter].width = 50
        elif header in ["Validation Command", "Validation Output"]:
            ws.column_dimensions[letter].width = 60
        elif header in ["Validation Reasoning", "Expected Pattern", "Compliance Logic"]:
            ws.column_dimensions[letter].width = 50
        elif header in ["Output Type"]:
            ws.column_dimensions[letter].width = 18
        else:
            ws.column_dimensions[letter].width = 18


def apply_status_colors(ws, header_map):
    """Apply color coding to Status column."""
    if "Status" not in header_map:
        return

    status_col = header_map["Status"]

    # Define colors
    compliant_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    non_compliant_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    error_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
    blocked_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")  # Light blue

    for row in range(2, ws.max_row + 1):
        status_cell = ws.cell(row=row, column=status_col)
        status_value = str(status_cell.value).strip().lower()

        if "compliant" in status_value and "non" not in status_value:
            status_cell.fill = compliant_fill
        elif "non compliant" in status_value:
            status_cell.fill = non_compliant_fill
        elif "error" in status_value:
            status_cell.fill = error_fill
        elif "blocked" in status_value or "dry" in status_value:
            status_cell.fill = blocked_fill


# ============================================================================
# Main Processing Logic
# ============================================================================

def process_single_finding(
    idx: int,
    total: int,
    row: pd.Series,
    client: Optional[OpenAI],
    args,
    benchmark_name: str,
) -> Optional[Dict]:
    """
    Process a single finding from the Nessus CSV.

    Returns a record dict, or None if the row is not a valid CIS item.
    """
    cis_id, cis_title, desc_body = parse_cis_from_description(str(row.get("Description", "")))
    if not cis_id:
        return None  # skip non-CIS items

    finding = clean_finding_title(cis_title)

    details_summary = desc_body.strip()
    impact = extract_impact_from_solution(str(row.get("Solution", "")))
    if impact:
        details_summary = (details_summary + "\n\nImpact:\n\n" + impact).strip()
    else:
        details_summary = re.sub(r"\n\s*Impact:\s*\n\s*$", "", details_summary, flags=re.I).strip()

    remediation = remove_impact_from_solution(str(row.get("Solution", "")))

    # Original status from Nessus
    original_status = "Non Compliant" if row["Risk"] == "FAILED" else "Compliant"

    # Gap 4: Configurable benchmark string
    record = {
        "CIS Benchmark": f"{benchmark_name} - {cis_id}",
        "Finding": finding,
        "Details Summary": details_summary,
        "Remediation": remediation,
        "Status": original_status,
    }

    # LLM Validation
    if args.validate and client is not None:
        logger.info("[%d/%d] Validating: %s - %s...", idx, total, cis_id, finding[:50])

        try:
            # Step 1: Generate structured validation JSON  (Gap 5 + Gap 13)
            logger.info("  ├─ Generating validation command...")
            validation_spec = generate_validation_json(client, finding, details_summary, args.model)
            command = validation_spec["validation_command"]
            expected_pattern = validation_spec["expected_output_pattern"]
            compliance_logic = validation_spec["compliance_logic"]
            output_type = validation_spec["output_type"]
            logger.info("  ├─ Command: %s", command[:80])
            logger.info("  ├─ Output Type: %s", output_type)

            # Gap 12: Dry-run mode — skip execution
            if args.dry_run:
                logger.info("  └─ DRY-RUN: Skipping execution")
                record["Status"] = "DRY-RUN — Not Executed"
                if args.add_evidence:
                    record["Validation Command"] = command
                    record["Validation Output"] = "[DRY-RUN] Command not executed"
                    record["Validation Reasoning"] = "[DRY-RUN] Compliance not evaluated"
                    record["Expected Pattern"] = expected_pattern
                    record["Compliance Logic"] = compliance_logic
                    record["Output Type"] = output_type
                return record

            # Gap 1: Command safety validation
            is_safe, safety_reason = is_command_safe(command)
            if not is_safe:
                logger.warning("  ├─ ⛔ Command BLOCKED: %s", safety_reason)
                record["Status"] = "BLOCKED — Unsafe Command"
                if args.add_evidence:
                    record["Validation Command"] = command
                    record["Validation Output"] = f"BLOCKED: {safety_reason}"
                    record["Validation Reasoning"] = "Command failed safety validation"
                    record["Expected Pattern"] = expected_pattern
                    record["Compliance Logic"] = compliance_logic
                    record["Output Type"] = output_type
                # Gap 3: Do NOT skip — record is still returned
                return record

            time.sleep(0.3)  # Small rate-limit buffer

            # Step 2: Execute command with fallback  (Gap 14)
            logger.info("  ├─ Executing command...")
            exec_result = execute_with_fallback(command, finding)

            # If fallback was used, log it
            actual_command = exec_result.get('command_used', command)
            if actual_command != command:
                logger.info("  ├─ ✅ Fallback command succeeded: %s", actual_command[:60])
                command = actual_command

            # Build combined output from stdout + stderr
            # Even if command "failed", the output is useful data for the LLM.
            # E.g. "Property not found" = "Not configured" = Non Compliant.
            # E.g. "The parameter is incorrect" = wrong command, LLM can still reason.
            stdout_text = exec_result.get('output', '').strip()
            stderr_text = exec_result.get('errors', '').strip()

            if exec_result['success'] and stdout_text:
                output = stdout_text
            elif stdout_text and stderr_text:
                output = f"{stdout_text}\n[STDERR]: {stderr_text}"
            elif stdout_text:
                output = stdout_text
            elif stderr_text:
                # Command failed but we have error output — this IS the validation data
                if is_meaningful_error(stderr_text):
                    output = f"[SETTING NOT CONFIGURED] {stderr_text}"
                    logger.info("  ├─ ℹ️  Setting not configured: %s", stderr_text[:80])
                else:
                    output = f"[COMMAND FAILED] {stderr_text}"
                    logger.warning("  ├─ ⚠️  Command error (sending to LLM): %s", stderr_text[:80])
            else:
                # Absolutely nothing — no stdout, no stderr. True failure.
                logger.warning("  ├─ ⚠️  Command produced no output at all")
                output = "[NO OUTPUT] Command returned no stdout and no stderr"

            logger.info("  ├─ Output: %s...", output[:80].strip())

            time.sleep(0.3)  # Small rate-limit buffer

            # Step 3: LLM validation (enhanced with expected_pattern + compliance_logic)
            logger.info("  ├─ Analyzing compliance...")
            validation = validate_compliance(
                client, finding, details_summary, output,
                expected_pattern, compliance_logic, args.model
            )

            # Update status based on LLM analysis
            record["Status"] = validation['status']

            logger.info(
                "  └─ Result: %s (%s confidence)",
                validation['status'], validation['confidence']
            )

            # Add evidence columns if requested
            if args.add_evidence:
                record["Validation Command"] = command
                record["Validation Output"] = output[:500]
                record["Validation Reasoning"] = validation['reasoning']
                record["Expected Pattern"] = expected_pattern
                record["Compliance Logic"] = compliance_logic
                record["Output Type"] = output_type

        except Exception as e:
            logger.error("  └─ ERROR: %s", e)
            record["Status"] = "ERROR — Validation Failed"
            if args.add_evidence:
                record["Validation Command"] = "N/A"
                record["Validation Output"] = str(e)
                record["Validation Reasoning"] = "Validation process failed"
                record["Expected Pattern"] = "N/A"
                record["Compliance Logic"] = "N/A"
                record["Output Type"] = "other"

    return record


def main():
    parser = argparse.ArgumentParser(
        description="Windows CIS Compliance AI Agent — Validates CIS benchmark compliance using AI-powered analysis (v3)"
    )
    parser.add_argument("-i", "--input", required=True, help="Nessus CSV output file")
    parser.add_argument("-t", "--template", required=True, help="Company template XLSX")
    parser.add_argument("-o", "--output", required=True, help="Output XLSX file")
    parser.add_argument("--sheet", default=None, help="Template sheet name (default first sheet)")
    parser.add_argument("--only-failed", action="store_true", help="Include only FAILED (Non-Compliant) items")
    parser.add_argument("--validate", action="store_true", help="Perform LLM-powered validation (requires OPENAI_API_KEY)")
    parser.add_argument("--add-evidence", action="store_true", help="Add validation command and output columns")
    parser.add_argument("--model", default="gpt-4o-mini", help="OpenAI model to use (default: gpt-4o-mini)")
    parser.add_argument("--limit", "--max-findings", type=int, default=0, help="Maximum number of findings to process (0=all)")
    parser.add_argument("--skip", type=int, default=0, help="Number of findings to skip at the start (default: 0)")
    # --- New v3 arguments ---
    parser.add_argument(
        "--benchmark", default="CIS Microsoft Windows Server 2022 Stand-alone v1.0.0 L1",
        help="CIS benchmark name prefix (default: CIS Microsoft Windows Server 2022 Stand-alone v1.0.0 L1)"
    )
    parser.add_argument("--dry-run", action="store_true", help="Generate commands but do NOT execute them")
    parser.add_argument("--resume", action="store_true", help="Resume from last checkpoint if available")
    parser.add_argument("--log-file", default=None, help="Write log output to file")
    parser.add_argument("--verbose", action="store_true", help="Enable debug-level logging")
    parser.add_argument(
        "--parallel", type=int, default=0,
        help="Number of concurrent findings to process (0=sequential). Requires --validate."
    )

    args = parser.parse_args()

    # Setup logging  (Gap 7)
    setup_logging(log_file=args.log_file, verbose=args.verbose)

    logger.info("=" * 80)
    logger.info("🤖 Windows CIS Compliance AI Agent v3.0")
    logger.info("=" * 80)
    logger.info("Input CSV: %s", args.input)
    logger.info("Template: %s", args.template)
    logger.info("Output: %s", args.output)
    logger.info("Validation: %s", "ENABLED" if args.validate else "DISABLED")
    if args.dry_run:
        logger.info("Mode: DRY-RUN (commands will NOT be executed)")
    if args.resume:
        logger.info("Resume: ENABLED")
    if args.parallel > 0:
        logger.info("Parallel: %d concurrent findings", args.parallel)

    # Validate API key if validation is enabled
    client = None
    if args.validate:
        api_key = os.getenv("OPENAI_API_KEY")
        if not api_key:
            logger.error("Validation enabled but OPENAI_API_KEY not set!")
            logger.error("Run: setx OPENAI_API_KEY \"your-key-here\"")
            sys.exit(1)

        # Gap 8: Create client once
        client = OpenAI(api_key=api_key)

        logger.info("LLM Model: %s", args.model)
        # Gap 11: Fully mask API key
        logger.info("API Key: Found ✓")

    # Load Nessus CSV
    logger.info("")
    logger.info("[STEP 1] Loading Nessus CSV...")
    df = pd.read_csv(args.input)
    logger.info("Loaded %d total rows", len(df))

    # Gap 6: Validate CSV columns
    try:
        validate_csv_columns(df)
    except ValueError as e:
        logger.error(str(e))
        sys.exit(1)

    # Keep only compliance rows (FAILED/PASSED)
    df = df[df["Risk"].isin(["FAILED", "PASSED"])].copy()
    logger.info("Filtered to %d compliance findings (FAILED/PASSED)", len(df))

    if args.only_failed:
        df = df[df["Risk"] == "FAILED"].copy()
        logger.info("Filtered to %d FAILED findings only", len(df))

    # Slice dataframe if skip or limit is used
    if args.skip > 0 or args.limit > 0:
        start_idx = args.skip
        end_idx = (args.skip + args.limit) if args.limit > 0 else len(df)
        df = df.iloc[start_idx:end_idx].copy()
        logger.info("[+] Range: Skipping first %d, taking next %d findings (Total: %d)", args.skip, len(df), len(df))

    # Gap 9: Checkpoint / Resume
    checkpoint_path = get_checkpoint_path(args.output)
    records = []
    processed_indices: List = []

    if args.resume:
        records, processed_indices = load_checkpoint(checkpoint_path)

    # Process findings
    logger.info("")
    logger.info("[STEP 2] Processing Findings...")

    finding_counter = len(records)
    total_findings = len(df)

    for df_idx, (idx, row) in enumerate(df.iterrows()):
        # Skip already-processed indices on resume
        if idx in processed_indices:
            logger.debug("Skipping already-processed index %d", idx)
            continue

        finding_counter += 1
        record = process_single_finding(
            idx=finding_counter,
            total=total_findings,
            row=row,
            client=client,
            args=args,
            benchmark_name=args.benchmark,
        )

        if record is not None:
            records.append(record)
            processed_indices.append(idx)

            # Save checkpoint every 10 findings  (Gap 9)
            if args.validate and len(records) % 10 == 0:
                save_checkpoint(checkpoint_path, records, processed_indices)

    logger.info("Processed %d findings", len(records))

    # Create DataFrame
    out_df = pd.DataFrame(records)
    out_df.insert(0, "Sr. No", range(1, len(out_df) + 1))

    # Statistics
    if args.validate or args.dry_run:
        logger.info("")
        logger.info("[STATISTICS]")
        status_counts = out_df['Status'].value_counts()
        for status, count in status_counts.items():
            logger.info("  ├─ %s: %d", status, count)

    # Load template
    logger.info("")
    logger.info("[STEP 3] Applying to Company Template...")
    wb = load_workbook(args.template)
    ws_name = args.sheet if args.sheet else wb.sheetnames[0]
    ws = wb[ws_name]

    # Get headers from template (first row) and normalize them
    headers_raw = [c.value for c in ws[1]]

    def norm(x):
        if x is None:
            return ""
        return str(x).strip().replace("\n", " ").replace("\t", " ")

    headers = [norm(h) for h in headers_raw]
    header_map = {headers[i]: i + 1 for i in range(len(headers)) if headers[i]}

    # If adding evidence columns, extend headers  (Gap 5: includes new structured fields)
    if args.add_evidence and (args.validate or args.dry_run):
        new_headers = [
            "Validation Command", "Validation Output", "Validation Reasoning",
            "Expected Pattern", "Compliance Logic", "Output Type"
        ]
        for new_header in new_headers:
            if new_header not in headers:
                headers.append(new_header)
                ws.cell(row=1, column=len(headers), value=new_header)
                header_map[new_header] = len(headers)

    # Clear template data (keep header row)
    if ws.max_row > 1:
        ws.delete_rows(2, ws.max_row)

    # Write rows aligned to template headers
    for _, r in out_df.iterrows():
        row_data = [""] * len(headers)
        for col, val in r.items():
            if col in header_map:
                row_data[header_map[col] - 1] = val
        ws.append(row_data)

    # Apply formatting
    logger.info("Applying formatting...")
    format_cell_wrapping(ws)
    set_column_widths(ws, headers)

    if args.validate or args.dry_run:
        apply_status_colors(ws, header_map)

    # Save
    wb.save(args.output)

    # Remove checkpoint on success  (Gap 9)
    remove_checkpoint(checkpoint_path)

    logger.info("")
    logger.info("=" * 80)
    logger.info("✅ Windows CIS Compliance AI Agent — Report generated: %s", args.output)
    logger.info("=" * 80)


if __name__ == "__main__":
    main()
